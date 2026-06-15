import os
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO
from collections import OrderedDict
from dateutil.relativedelta import relativedelta

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, send_file, abort)
from flask_login import (LoginManager, login_user, logout_user, login_required,
                         current_user)
from sqlalchemy import func, and_, or_
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from config import Config
from models import (db, User, Supplier, Order, Inspection, PaymentRequest, Receipt)


def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)

    db.init_app(app)
    login_manager = LoginManager(app)
    login_manager.login_view = 'login'
    login_manager.login_message = '请先登录'

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    def role_required(*roles):
        def decorator(f):
            @wraps(f)
            def decorated_function(*args, **kwargs):
                if not current_user.is_authenticated:
                    return redirect(url_for('login'))
                if current_user.role not in roles:
                    flash('您没有权限执行此操作', 'danger')
                    return redirect(url_for('dashboard'))
                return f(*args, **kwargs)
            return decorated_function
        return decorator

    def generate_order_no():
        today = date.today().strftime('%Y%m%d')
        prefix = f'WW{today}'
        last_order = Order.query.filter(Order.order_no.like(f'{prefix}%')).order_by(Order.order_no.desc()).first()
        seq = int(last_order.order_no[-4:]) + 1 if last_order else 1
        return f'{prefix}{seq:04d}'

    def generate_payment_no():
        today = date.today().strftime('%Y%m%d')
        prefix = f'FK{today}'
        last = PaymentRequest.query.filter(PaymentRequest.request_no.like(f'{prefix}%')).order_by(PaymentRequest.request_no.desc()).first()
        seq = int(last.request_no[-4:]) + 1 if last else 1
        return f'{prefix}{seq:04d}'

    def generate_receipt_no():
        today = date.today().strftime('%Y%m%d')
        prefix = f'SH{today}'
        last = Receipt.query.filter(Receipt.receipt_no.like(f'{prefix}%')).order_by(Receipt.receipt_no.desc()).first()
        seq = int(last.receipt_no[-4:]) + 1 if last else 1
        return f'{prefix}{seq:04d}'

    def generate_inspection_no():
        today = date.today().strftime('%Y%m%d')
        prefix = f'ZJ{today}'
        last = Inspection.query.filter(Inspection.inspection_no.like(f'{prefix}%')).order_by(Inspection.inspection_no.desc()).first()
        seq = int(last.inspection_no[-4:]) + 1 if last else 1
        return f'{prefix}{seq:04d}'

    def apply_order_filters(query, params):
        if current_user.role == 'supplier':
            query = query.filter_by(supplier_id=current_user.supplier_id)
        supplier_id = params.get('supplier_id', type=int)
        if supplier_id and current_user.role == 'enterprise':
            query = query.filter_by(supplier_id=supplier_id)
        status = params.get('status', '')
        if status:
            query = query.filter_by(status=status)
        start_date = params.get('start_date', '')
        if start_date:
            try:
                sd = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(Order.created_at >= sd)
            except: pass
        end_date = params.get('end_date', '')
        if end_date:
            try:
                ed = datetime.strptime(end_date, '%Y-%m-%d')
                ed = ed.replace(hour=23, minute=59, second=59)
                query = query.filter(Order.created_at <= ed)
            except: pass
        return query

    @app.context_processor
    def inject_common_vars():
        warnings = []
        if current_user.is_authenticated and current_user.role == 'enterprise':
            for s in Supplier.query.all():
                pending = s.get_pending_quantity()
                if pending > s.monthly_capacity:
                    warnings.append({
                        'supplier': s.name,
                        'supplier_id': s.id,
                        'pending': pending,
                        'capacity': s.monthly_capacity,
                        'overload': pending - s.monthly_capacity
                    })
        return dict(capacity_warnings=warnings, date=date)

    # ======== 认证路由 ========

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                flash('登录成功', 'success')
                return redirect(url_for('dashboard'))
            flash('用户名或密码错误', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        flash('已退出登录', 'info')
        return redirect(url_for('login'))

    # ======== 仪表板 ========

    @app.route('/')
    @login_required
    def dashboard():
        stats = {}
        recent_orders_query = Order.query

        if current_user.role == 'enterprise':
            stats['total_orders'] = Order.query.count()
            stats['pending_orders'] = Order.query.filter(Order.status.in_(['已下单', '已接单', '生产中'])).count()
            stats['shipping_orders'] = Order.query.filter(Order.status.in_(['已发货', '部分到货'])).count()
            stats['inspection_orders'] = Order.query.filter(Order.status.in_(['已到货', '质检中'])).count()
            stats['completed_orders'] = Order.query.filter(Order.status.in_(['质检完成', '已完成'])).count()
            stats['total_payable'] = 0
            for o in Order.query.filter(Order.status.in_(['质检完成', '已完成'])).all():
                stats['total_payable'] += o.get_payable_amount()
            stats['pending_payment'] = db.session.query(func.sum(PaymentRequest.amount)).filter(PaymentRequest.status != '已付款').scalar() or 0
        else:
            sid = current_user.supplier_id
            stats['total_orders'] = Order.query.filter_by(supplier_id=sid).count()
            stats['new_orders'] = Order.query.filter_by(supplier_id=sid, status='已下单').count()
            stats['in_production'] = Order.query.filter_by(supplier_id=sid, status='生产中').count()
            stats['shipped'] = Order.query.filter(Order.supplier_id == sid, Order.status.in_(['已发货', '部分到货'])).count()
            stats['inspection'] = Order.query.filter(Order.supplier_id == sid, Order.status.in_(['已到货', '质检中'])).count()
            stats['completed'] = Order.query.filter(Order.supplier_id == sid, Order.status.in_(['质检完成', '已完成'])).count()
            pending_amount = 0
            for o in Order.query.filter(Order.supplier_id == sid, Order.status == '质检完成').all():
                if not o.payment_requests:
                    pending_amount += o.get_payable_amount()
            stats['pending_amount'] = pending_amount
            recent_orders_query = recent_orders_query.filter_by(supplier_id=sid)

        recent_orders = recent_orders_query.order_by(Order.created_at.desc()).limit(10).all()
        return render_template('dashboard.html', stats=stats, recent_orders=recent_orders)

    # ======== 外协厂管理 ========

    @app.route('/suppliers')
    @login_required
    @role_required('enterprise')
    def list_suppliers():
        suppliers = Supplier.query.all()
        supplier_stats = []
        for s in suppliers:
            pending = s.get_pending_quantity()
            supplier_stats.append({
                'supplier': s,
                'pending_qty': pending,
                'capacity_ratio': f'{(pending / s.monthly_capacity * 100):.1f}%' if s.monthly_capacity > 0 else 'N/A',
                'overloaded': pending > s.monthly_capacity,
                'pending_apply': s.get_pending_apply_amount(),
                'pending_approve': s.get_pending_approve_amount(),
                'pending_pay': s.get_pending_pay_amount(),
                'unsettled': s.get_unsettled_amount()
            })
        return render_template('suppliers/list.html', supplier_stats=supplier_stats)

    @app.route('/suppliers/<int:supplier_id>/edit-capacity', methods=['POST'])
    @login_required
    @role_required('enterprise')
    def edit_supplier_capacity(supplier_id):
        s = Supplier.query.get_or_404(supplier_id)
        s.monthly_capacity = int(request.form.get('monthly_capacity', s.monthly_capacity))
        s.contact = request.form.get('contact', s.contact)
        s.phone = request.form.get('phone', s.phone)
        s.address = request.form.get('address', s.address)
        db.session.commit()
        flash('外协厂信息已更新', 'success')
        return redirect(url_for('list_suppliers'))

    @app.route('/api/suppliers/<int:supplier_id>/capacity')
    @login_required
    @role_required('enterprise')
    def get_supplier_capacity(supplier_id):
        s = Supplier.query.get_or_404(supplier_id)
        pending = s.get_pending_quantity()
        return jsonify({
            'supplier': s.name,
            'pending': pending,
            'capacity': s.monthly_capacity,
            'overloaded': pending > s.monthly_capacity,
            'available': max(0, s.monthly_capacity - pending)
        })

    # ======== 结算看板 ========

    @app.route('/settlement-dashboard')
    @login_required
    @role_required('enterprise')
    def settlement_dashboard():
        supplier_filter = request.args.get('supplier_id', type=int)
        suppliers = Supplier.query.all()
        settlement_data = []
        total_pending_apply = 0
        total_pending_approve = 0
        total_pending_pay = 0
        total_unsettled = 0

        for s in suppliers:
            if supplier_filter and s.id != supplier_filter:
                continue
            pending_apply = s.get_pending_apply_amount()
            pending_approve = s.get_pending_approve_amount()
            pending_pay = s.get_pending_pay_amount()
            unsettled = s.get_unsettled_amount()
            total_pending_apply += pending_apply
            total_pending_approve += pending_approve
            total_pending_pay += pending_pay
            total_unsettled += unsettled

            orders_to_apply = []
            for o in Order.query.filter(
                Order.supplier_id == s.id,
                Order.status == '质检完成'
            ).all():
                if not o.payment_requests:
                    orders_to_apply.append(o)

            settlement_data.append({
                'supplier': s,
                'pending_apply': pending_apply,
                'pending_approve': pending_approve,
                'pending_pay': pending_pay,
                'unsettled': unsettled,
                'orders_to_apply': orders_to_apply,
                'pending_payments': PaymentRequest.query.filter_by(
                    supplier_id=s.id, status='待审批'
                ).order_by(PaymentRequest.created_at.desc()).all(),
                'approved_payments': PaymentRequest.query.filter_by(
                    supplier_id=s.id, status='已审批'
                ).order_by(PaymentRequest.created_at.desc()).all(),
            })

        totals = {
            'pending_apply': total_pending_apply,
            'pending_approve': total_pending_approve,
            'pending_pay': total_pending_pay,
            'unsettled': total_unsettled
        }
        return render_template('settlement/dashboard.html',
                               settlement_data=settlement_data,
                               suppliers=suppliers,
                               supplier_filter=supplier_filter,
                               totals=totals)

    # ======== 订单管理 ========

    @app.route('/orders')
    @login_required
    def list_orders():
        query = Order.query
        query = apply_order_filters(query, request.args)
        orders = query.order_by(Order.created_at.desc()).all()
        suppliers = Supplier.query.all() if current_user.role == 'enterprise' else []
        statuses = ['已下单', '已接单', '生产中', '已发货', '部分到货', '已到货', '质检中', '质检完成', '已完成']
        return render_template('orders/list.html', orders=orders, suppliers=suppliers,
                               status_filter=request.args.get('status', ''),
                               supplier_filter=request.args.get('supplier_id', type=int),
                               start_date=request.args.get('start_date', ''),
                               end_date=request.args.get('end_date', ''),
                               statuses=statuses)

    @app.route('/orders/new', methods=['GET', 'POST'])
    @login_required
    @role_required('enterprise')
    def create_order():
        suppliers = Supplier.query.all()
        if request.method == 'POST':
            supplier_id = int(request.form['supplier_id'])
            supplier = Supplier.query.get(supplier_id)
            quantity = int(request.form['quantity'])
            pending = supplier.get_pending_quantity()
            if pending + quantity > supplier.monthly_capacity:
                flash(f'警告: {supplier.name} 当前未完成{pending}件，月产能{supplier.monthly_capacity}件，'
                      f'添加{quantity}件后将超出{pending + quantity - supplier.monthly_capacity}件！', 'warning')
            order = Order(
                order_no=generate_order_no(),
                supplier_id=supplier_id,
                drawing_no=request.form['drawing_no'],
                part_name=request.form.get('part_name', ''),
                quantity=quantity,
                unit_price=float(request.form['unit_price']),
                agreed_delivery_date=datetime.strptime(request.form['agreed_delivery_date'], '%Y-%m-%d').date(),
                remark=request.form.get('remark', '')
            )
            db.session.add(order)
            db.session.commit()
            flash(f'订单 {order.order_no} 创建成功', 'success')
            return redirect(url_for('view_order', order_id=order.id))
        return render_template('orders/new.html', suppliers=suppliers)

    @app.route('/orders/<int:order_id>')
    @login_required
    def view_order(order_id):
        order = Order.query.get_or_404(order_id)
        if current_user.role == 'supplier' and order.supplier_id != current_user.supplier_id:
            abort(403)
        pending_receipts = [r for r in order.receipts if r.get_pending_inspection() > 0]
        return render_template('orders/detail.html', order=order, pending_receipts=pending_receipts)

    @app.route('/orders/<int:order_id>/update-status', methods=['POST'])
    @login_required
    def update_order_status(order_id):
        order = Order.query.get_or_404(order_id)
        new_status = request.form['status']
        now = datetime.now()
        valid_transitions = {
            'enterprise': {'已发货': '已到货', '部分到货': '已到货'},
            'supplier': {'已下单': '已接单', '已接单': '生产中', '生产中': '已发货'}
        }
        if current_user.role == 'supplier' and order.supplier_id != current_user.supplier_id:
            abort(403)
        allowed = valid_transitions.get(current_user.role, {})
        if order.status not in allowed or allowed[order.status] != new_status:
            flash('非法的状态转换', 'danger')
            return redirect(url_for('view_order', order_id=order.id))
        order.status = new_status
        if new_status == '已接单':
            order.accepted_at = now
        elif new_status == '生产中':
            order.production_at = now
        elif new_status == '已发货':
            order.shipped_at = now
        elif new_status == '已到货':
            order.last_arrived_at = now
        db.session.commit()
        flash(f'订单状态已更新为: {new_status}', 'success')
        return redirect(url_for('view_order', order_id=order.id))

    @app.route('/orders/<int:order_id>/edit', methods=['GET', 'POST'])
    @login_required
    @role_required('enterprise')
    def edit_order(order_id):
        order = Order.query.get_or_404(order_id)
        if order.status not in ['已下单']:
            flash('仅"已下单"状态的订单可编辑', 'warning')
            return redirect(url_for('view_order', order_id=order.id))
        suppliers = Supplier.query.all()
        if request.method == 'POST':
            order.supplier_id = int(request.form['supplier_id'])
            order.drawing_no = request.form['drawing_no']
            order.part_name = request.form.get('part_name', '')
            order.quantity = int(request.form['quantity'])
            order.unit_price = float(request.form['unit_price'])
            order.agreed_delivery_date = datetime.strptime(request.form['agreed_delivery_date'], '%Y-%m-%d').date()
            order.remark = request.form.get('remark', '')
            db.session.commit()
            flash('订单已更新', 'success')
            return redirect(url_for('view_order', order_id=order.id))
        return render_template('orders/edit.html', order=order, suppliers=suppliers)

    @app.route('/orders/<int:order_id>/delete', methods=['POST'])
    @login_required
    @role_required('enterprise')
    def delete_order(order_id):
        order = Order.query.get_or_404(order_id)
        if order.status not in ['已下单']:
            flash('仅"已下单"状态的订单可删除', 'danger')
            return redirect(url_for('view_order', order_id=order.id))
        order_no = order.order_no
        db.session.delete(order)
        db.session.commit()
        flash(f'订单 {order_no} 已删除', 'info')
        return redirect(url_for('list_orders'))

    # ======== 收货单管理 ========

    @app.route('/orders/<int:order_id>/receipt', methods=['GET', 'POST'])
    @login_required
    @role_required('enterprise')
    def create_receipt(order_id):
        order = Order.query.get_or_404(order_id)
        if not order.can_create_receipt():
            flash('当前状态不支持收货，或订单数量已全部收齐', 'warning')
            return redirect(url_for('view_order', order_id=order.id))
        if request.method == 'POST':
            qty = int(request.form['received_quantity'])
            remaining = order.get_remaining_quantity()
            if qty <= 0 or qty > remaining:
                flash(f'收货数量必须大于0且不超过剩余数量({remaining})', 'danger')
                return redirect(url_for('create_receipt', order_id=order.id))
            receipt = Receipt(
                receipt_no=generate_receipt_no(),
                order_id=order.id,
                supplier_id=order.supplier_id,
                received_quantity=qty,
                waybill_no=request.form.get('waybill_no', ''),
                receiver=request.form.get('receiver', current_user.username),
                remark=request.form.get('remark', '')
            )
            db.session.add(receipt)
            order.update_status_after_receipt()
            db.session.commit()
            flash(f'收货单 {receipt.receipt_no} 创建成功，已收货{qty}件', 'success')
            if order.status == '部分到货':
                return redirect(url_for('view_order', order_id=order.id))
            return redirect(url_for('view_order', order_id=order.id))
        return render_template('orders/receipt.html', order=order)

    # ======== 质检管理 ========

    @app.route('/receipts/<int:receipt_id>/inspect', methods=['GET', 'POST'])
    @login_required
    @role_required('enterprise')
    def inspect_receipt(receipt_id):
        receipt = Receipt.query.get_or_404(receipt_id)
        order = receipt.order
        pending = receipt.get_pending_inspection()
        if pending <= 0:
            flash('该收货单已全部质检完成', 'info')
            return redirect(url_for('view_order', order_id=order.id))
        if request.method == 'POST':
            qualified = int(request.form['qualified_quantity'])
            unqualified = int(request.form['unqualified_quantity'])
            if qualified + unqualified != pending:
                flash(f'合格数量+不合格数量必须等于待检数量({pending})', 'danger')
                return redirect(url_for('inspect_receipt', receipt_id=receipt.id))
            if qualified < 0 or unqualified < 0:
                flash('数量不能为负数', 'danger')
                return redirect(url_for('inspect_receipt', receipt_id=receipt.id))
            defect_reasons = request.form.getlist('defect_reasons')
            defect_detail = request.form.get('defect_detail', '').strip()
            if unqualified > 0 and not defect_reasons:
                flash('不合格数量大于0时，必须勾选不合格原因', 'danger')
                return redirect(url_for('inspect_receipt', receipt_id=receipt.id))
            inspection = Inspection(
                inspection_no=generate_inspection_no(),
                order_id=order.id,
                receipt_id=receipt.id,
                qualified_quantity=qualified,
                unqualified_quantity=unqualified,
                defect_reasons=','.join(defect_reasons) if defect_reasons else '',
                defect_detail=defect_detail,
                inspector=current_user.username,
                remark=request.form.get('remark', '')
            )
            db.session.add(inspection)
            order.update_status_after_inspection()
            db.session.commit()
            flash(f'质检完成：合格{qualified}件，不合格{unqualified}件，本次应付¥{inspection.get_inspection_amount():,.2f}', 'success')
            return redirect(url_for('view_order', order_id=order.id))
        defect_options = ['尺寸超差', '表面划伤', '变形', '材质不符', '热处理不合格', '其他']
        return render_template('orders/inspect.html', order=order, receipt=receipt,
                               pending=pending, defect_options=defect_options)

    # ======== 付款申请 ========

    @app.route('/payments')
    @login_required
    @role_required('enterprise')
    def list_payments():
        status_filter = request.args.get('status', '')
        supplier_filter = request.args.get('supplier_id', type=int)
        query = PaymentRequest.query
        if status_filter:
            query = query.filter_by(status=status_filter)
        if supplier_filter:
            query = query.filter_by(supplier_id=supplier_filter)
        payments = query.order_by(PaymentRequest.created_at.desc()).all()
        suppliers = Supplier.query.all()
        return render_template('payments/list.html', payments=payments, status_filter=status_filter,
                               supplier_filter=supplier_filter, suppliers=suppliers)

    @app.route('/payments/new/<int:order_id>', methods=['GET', 'POST'])
    @login_required
    @role_required('enterprise')
    def create_payment(order_id):
        order = Order.query.get_or_404(order_id)
        if order.status != '质检完成':
            flash('仅"质检完成"状态的订单可创建付款申请', 'warning')
            return redirect(url_for('view_order', order_id=order.id))
        existing = PaymentRequest.query.filter_by(order_id=order.id).first()
        if existing:
            flash('该订单已有付款申请', 'info')
            return redirect(url_for('view_payment', payment_id=existing.id))
        if request.method == 'POST':
            payment = PaymentRequest(
                request_no=generate_payment_no(),
                order_id=order.id,
                supplier_id=order.supplier_id,
                amount=order.get_payable_amount(),
                applicant=current_user.username,
                remark=request.form.get('remark', '')
            )
            db.session.add(payment)
            db.session.commit()
            flash(f'付款申请 {payment.request_no} 创建成功', 'success')
            return redirect(url_for('view_payment', payment_id=payment.id))
        return render_template('payments/new.html', order=order)

    @app.route('/payments/<int:payment_id>')
    @login_required
    @role_required('enterprise')
    def view_payment(payment_id):
        payment = PaymentRequest.query.get_or_404(payment_id)
        return render_template('payments/detail.html', payment=payment)

    @app.route('/payments/<int:payment_id>/approve', methods=['POST'])
    @login_required
    @role_required('enterprise')
    def approve_payment(payment_id):
        payment = PaymentRequest.query.get_or_404(payment_id)
        if payment.status != '待审批':
            flash('仅待审批的申请可审批', 'warning')
            return redirect(url_for('view_payment', payment_id=payment.id))
        payment.status = '已审批'
        payment.approver = current_user.username
        payment.approved_at = datetime.now()
        db.session.commit()
        flash('付款申请已审批', 'success')
        return redirect(url_for('view_payment', payment_id=payment.id))

    @app.route('/payments/<int:payment_id>/pay', methods=['POST'])
    @login_required
    @role_required('enterprise')
    def pay_payment(payment_id):
        payment = PaymentRequest.query.get_or_404(payment_id)
        if payment.status != '已审批':
            flash('仅已审批的申请可付款', 'warning')
            return redirect(url_for('view_payment', payment_id=payment.id))
        payment.status = '已付款'
        payment.paid_at = datetime.now()
        order = payment.order
        order.status = '已完成'
        order.completed_at = datetime.now()
        db.session.commit()
        flash('付款完成，订单已关闭', 'success')
        return redirect(url_for('view_payment', payment_id=payment.id))

    # ======== 绩效报告 ========

    @app.route('/reports/performance')
    @login_required
    @role_required('enterprise')
    def performance_report():
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        supplier_filter = request.args.get('supplier_id', type=int)
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

        suppliers = Supplier.query.all()
        report_data = []
        for s in suppliers:
            if supplier_filter and s.id != supplier_filter:
                continue
            order_query = Order.query.filter_by(supplier_id=s.id).filter(Order.status != '已下单')
            if start_date:
                order_query = order_query.filter(Order.created_at >= datetime.combine(start_date, datetime.min.time()))
            if end_date:
                order_query = order_query.filter(Order.created_at <= datetime.combine(end_date, datetime.max.time()))
            orders = order_query.all()

            completed = [o for o in orders if o.status in ['质检完成', '已完成']]
            total_completed = len(completed)
            on_time_count = sum(1 for o in completed if o.is_on_time() is True)
            on_time_rate = (on_time_count / total_completed * 100) if total_completed > 0 else 0

            inspected_count = 0
            first_pass_count = 0
            total_processing_days = 0
            for o in completed:
                if o.inspections:
                    inspected_count += 1
                    if o.is_first_pass() is True:
                        first_pass_count += 1
                    days = o.get_processing_days()
                    total_processing_days += days
            first_pass_rate = (first_pass_count / inspected_count * 100) if inspected_count > 0 else 0
            avg_cycle = (total_processing_days / total_completed) if total_completed > 0 else 0

            total_value = sum(o.get_payable_amount() for o in completed)

            report_data.append({
                'supplier': s,
                'total_orders': len(orders),
                'completed_orders': total_completed,
                'on_time_count': on_time_count,
                'on_time_rate': on_time_rate,
                'first_pass_count': first_pass_count,
                'inspected_count': inspected_count,
                'first_pass_rate': first_pass_rate,
                'avg_cycle': avg_cycle,
                'total_value': total_value,
                'monthly_trend': s.get_monthly_performance(6),
                'cycle_max': max([avg_cycle, 20] + [m.get('avg_cycle', 0) for m in s.get_monthly_performance(6).values() if m.get('completed', 0) > 0])
            })

        all_suppliers = Supplier.query.all()
        month_labels = []
        today = date.today()
        for i in range(5, -1, -1):
            m_date = today - relativedelta(months=i)
            month_labels.append(f'{m_date.year}-{m_date.month:02d}')

        return render_template('reports/performance.html',
                               report_data=report_data,
                               suppliers=all_suppliers,
                               supplier_filter=supplier_filter,
                               start_date=start_date_str,
                               end_date=end_date_str,
                               month_labels=month_labels)

    # ======== Excel 导出 ========

    def style_header(ws, row, cols):
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_body(ws, start_row, end_row, cols):
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for r in range(start_row, end_row + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).border = thin_border
                ws.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)

    @app.route('/export/unsettled-orders')
    @login_required
    @role_required('enterprise')
    def export_unsettled_orders():
        unsettled_statuses = ['已下单', '已接单', '生产中', '已发货', '部分到货', '已到货', '质检中', '质检完成']
        query = Order.query.filter(Order.status.in_(unsettled_statuses))
        query = apply_order_filters(query, request.args)
        orders = query.order_by(Order.created_at.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = '未结订单清单'

        headers = ['订单号', '外协厂', '零件图号', '零件名称', '订单数量', '已收货', '已检验',
                   '合格数', '不合格数', '单价(元)', '总金额(元)', '应付金额(元)',
                   '约定交货日期', '当前状态', '下单时间', '已接单时间', '已发货时间',
                   '首次到货', '最后到货', '备注']
        ws.append(headers)
        style_header(ws, 1, len(headers))

        row = 2
        total_amt = 0
        total_payable = 0
        for o in orders:
            payable = o.get_payable_amount()
            total_amt += o.get_total_amount()
            total_payable += payable
            ws.append([
                o.order_no,
                o.supplier.name,
                o.drawing_no,
                o.part_name or '',
                o.quantity,
                o.get_total_received(),
                o.get_total_inspected(),
                o.get_total_qualified(),
                o.get_total_unqualified(),
                o.unit_price,
                o.get_total_amount(),
                payable,
                o.agreed_delivery_date.strftime('%Y-%m-%d'),
                o.status,
                o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else '',
                o.accepted_at.strftime('%Y-%m-%d %H:%M') if o.accepted_at else '',
                o.shipped_at.strftime('%Y-%m-%d %H:%M') if o.shipped_at else '',
                o.first_arrived_at.strftime('%Y-%m-%d %H:%M') if o.first_arrived_at else '',
                o.last_arrived_at.strftime('%Y-%m-%d %H:%M') if o.last_arrived_at else '',
                o.remark or ''
            ])
            row += 1

        ws.append(['', '', '', '', '', '', '', '', '', '合计', total_amt, total_payable,
                   '', '', '', '', '', '', '', ''])
        ws.cell(row=row, column=10).font = Font(bold=True)
        ws.cell(row=row, column=11).font = Font(bold=True)
        ws.cell(row=row, column=12).font = Font(bold=True)
        style_body(ws, 2, row, len(headers))

        col_widths = [18, 18, 14, 16, 10, 10, 10, 10, 10, 10, 12, 12, 14, 10, 16, 16, 16, 16, 16, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'未结订单清单_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/export/payables')
    @login_required
    @role_required('enterprise')
    def export_payables():
        supplier_filter = request.args.get('supplier_id', type=int)
        status_filter = request.args.get('status', '')
        payables = []

        orders_query = Order.query.filter(Order.status.in_(['质检完成', '已完成']))
        if supplier_filter:
            orders_query = orders_query.filter_by(supplier_id=supplier_filter)

        for o in orders_query.order_by(Order.last_inspected_at.desc()).all():
            if o.inspections:
                for pr in o.payment_requests:
                    if status_filter and pr.status != status_filter:
                        continue
                    payables.append({'order': o, 'payment': pr})
                if not o.payment_requests:
                    if not status_filter or status_filter == '未申请':
                        payables.append({'order': o, 'payment': None})

        wb = Workbook()
        ws = wb.active
        ws.title = '应付款明细'

        headers = ['订单号', '外协厂', '零件图号', '零件名称', '订单数量', '合格数量', '不合格数量',
                   '单价(元)', '应付金额(元)', '已收货', '已检验', '付款申请号', '申请状态',
                   '申请时间', '审批时间', '付款时间', '申请人', '审批人', '备注']
        ws.append(headers)
        style_header(ws, 1, len(headers))

        row = 2
        total_amount = 0
        for item in payables:
            o = item['order']
            pr = item['payment']
            payable = o.get_payable_amount()
            total_amount += payable
            ws.append([
                o.order_no,
                o.supplier.name,
                o.drawing_no,
                o.part_name or '',
                o.quantity,
                o.get_total_qualified(),
                o.get_total_unqualified(),
                o.unit_price,
                payable,
                o.get_total_received(),
                o.get_total_inspected(),
                pr.request_no if pr else '-',
                pr.status if pr else '未申请',
                pr.created_at.strftime('%Y-%m-%d %H:%M') if pr and pr.created_at else '',
                pr.approved_at.strftime('%Y-%m-%d %H:%M') if pr and pr.approved_at else '',
                pr.paid_at.strftime('%Y-%m-%d %H:%M') if pr and pr.paid_at else '',
                pr.applicant if pr else '',
                pr.approver if pr else '',
                pr.remark if pr else (o.remark or '')
            ])
            row += 1

        ws.append(['', '', '', '', '', '', '', '合计:', total_amount, '', '', '', '', '', '', '', '', '', '', ''])
        ws.cell(row=row, column=8).font = Font(bold=True)
        ws.cell(row=row, column=9).font = Font(bold=True)
        style_body(ws, 2, row, len(headers))

        col_widths = [18, 18, 14, 16, 10, 10, 10, 10, 12, 10, 10, 18, 10, 16, 16, 16, 10, 10, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'应付款明细_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ======== 月度付款分析 ========

    @app.route('/reports/payment-analysis')
    @login_required
    @role_required('enterprise')
    def payment_analysis():
        supplier_filter = request.args.get('supplier_id', type=int)
        suppliers = Supplier.query.all()
        analysis_data = []
        month_labels = []
        today = date.today()
        for i in range(5, -1, -1):
            m_date = today - relativedelta(months=i)
            month_labels.append(f'{m_date.year}-{m_date.month:02d}')

        for s in suppliers:
            if supplier_filter and s.id != supplier_filter:
                continue
            monthly = s.get_monthly_payment_summary(6)
            total_paid = sum(m['paid'] for m in monthly.values())
            total_unpaid = sum(m['unpaid'] for m in monthly.values())
            total_overdue = sum(m['overdue_unpaid'] for m in monthly.values())
            amt_max = max([10000, total_paid + total_unpaid] + [m['paid'] + m['unpaid'] for m in monthly.values()])
            analysis_data.append({
                'supplier': s,
                'monthly': monthly,
                'total_paid': total_paid,
                'total_unpaid': total_unpaid,
                'total_overdue': total_overdue,
                'amt_max': amt_max
            })

        totals = {
            'paid': sum(d['total_paid'] for d in analysis_data),
            'unpaid': sum(d['total_unpaid'] for d in analysis_data),
            'overdue': sum(d['total_overdue'] for d in analysis_data)
        }
        return render_template('reports/payment_analysis.html',
                               analysis_data=analysis_data,
                               suppliers=suppliers,
                               supplier_filter=supplier_filter,
                               month_labels=month_labels,
                               totals=totals)

    @app.route('/export/payment-monthly')
    @login_required
    @role_required('enterprise')
    def export_payment_monthly():
        supplier_filter = request.args.get('supplier_id', type=int)
        suppliers = Supplier.query.all()
        month_labels = []
        today = date.today()
        for i in range(5, -1, -1):
            m_date = today - relativedelta(months=i)
            month_labels.append(f'{m_date.year}-{m_date.month:02d}')

        wb = Workbook()
        for s in suppliers:
            if supplier_filter and s.id != supplier_filter:
                continue
            ws = wb.create_sheet(title=s.name[:20])
            headers = ['月份', '已付金额', '未付金额', '逾期未付', '合计']
            ws.append(headers)
            style_header(ws, 1, len(headers))
            monthly = s.get_monthly_payment_summary(6)
            row = 2
            total_paid = 0
            total_unpaid = 0
            total_overdue = 0
            for month_key in month_labels:
                mdata = monthly.get(month_key, {'paid': 0, 'unpaid': 0, 'overdue_unpaid': 0})
                paid = mdata['paid']
                unpaid = mdata['unpaid']
                overdue = mdata['overdue_unpaid']
                total_paid += paid
                total_unpaid += unpaid
                total_overdue += overdue
                ws.append([month_key, paid, unpaid, overdue, paid + unpaid])
                row += 1
            ws.append(['合计', total_paid, total_unpaid, total_overdue, total_paid + total_unpaid])
            ws.cell(row=row, column=1).font = Font(bold=True)
            style_body(ws, 2, row, len(headers))
            for i, w in enumerate([14, 14, 14, 14, 14], 1):
                ws.column_dimensions[chr(64 + i)].width = w

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'月度付款分析_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ======== 初始化数据库 ========

    with app.app_context():
        db.create_all()
        if not User.query.first():
            s1 = Supplier(name='精诚精密加工厂', contact='张经理', phone='13800138001',
                          address='江苏省苏州市工业园区XX路1号', monthly_capacity=2000)
            s2 = Supplier(name='宏远机械制造有限公司', contact='李总', phone='13800138002',
                          address='浙江省宁波市鄞州区XX路88号', monthly_capacity=3000)
            s3 = Supplier(name='顺达五金配件厂', contact='王厂长', phone='13800138003',
                          address='广东省东莞市长安镇XX路168号', monthly_capacity=1500)
            db.session.add_all([s1, s2, s3])
            db.session.flush()

            u_admin = User(username='admin', role='enterprise')
            u_admin.set_password('admin123')
            u_s1 = User(username='jingcheng', role='supplier', supplier_id=s1.id)
            u_s1.set_password('jc123456')
            u_s2 = User(username='hongyuan', role='supplier', supplier_id=s2.id)
            u_s2.set_password('hy123456')
            u_s3 = User(username='shunda', role='supplier', supplier_id=s3.id)
            u_s3.set_password('sd123456')
            db.session.add_all([u_admin, u_s1, u_s2, u_s3])

            today = date.today()
            demo_orders = [
                {'s': s1, 'drawing': 'DJ-2024-A001', 'name': '法兰盘', 'qty': 200, 'price': 45.5, 'days': 7, 'status': '已下单'},
                {'s': s2, 'drawing': 'DJ-2024-B002', 'name': '轴承座', 'qty': 500, 'price': 32.0, 'days': 10, 'status': '已接单'},
                {'s': s3, 'drawing': 'DJ-2024-C003', 'name': '连接支架', 'qty': 300, 'price': 28.5, 'days': 5, 'status': '生产中'},
                {'s': s1, 'drawing': 'DJ-2024-A004', 'name': '齿轮轴', 'qty': 150, 'price': 68.0, 'days': 14, 'status': '已发货'},
            ]
            for i, do in enumerate(demo_orders):
                created = today - timedelta(days=i + 1)
                o = Order(
                    order_no=f'WW{created.strftime("%Y%m%d")}{i + 1:04d}',
                    supplier_id=do['s'].id,
                    drawing_no=do['drawing'],
                    part_name=do['name'],
                    quantity=do['qty'],
                    unit_price=do['price'],
                    agreed_delivery_date=today + timedelta(days=do['days']),
                    status=do['status'],
                    created_at=datetime.combine(created, datetime.min.time()) + timedelta(hours=9 + i)
                )
                if do['status'] in ['已接单', '生产中', '已发货']:
                    o.accepted_at = o.created_at + timedelta(hours=2)
                if do['status'] in ['生产中', '已发货']:
                    o.production_at = o.accepted_at + timedelta(hours=4)
                if do['status'] == '已发货':
                    o.shipped_at = o.production_at + timedelta(days=2)
                db.session.add(o)

            db.session.commit()
            print('='*60)
            print('数据库初始化完成！默认账号:')
            print('企业管理员: admin / admin123')
            print('外协厂-精诚: jingcheng / jc123456')
            print('外协厂-宏远: hongyuan / hy123456')
            print('外协厂-顺达: shunda / sd123456')
            print('='*60)

    return app


if __name__ == '__main__':
    app = create_app()
    app.run(debug=True, host='0.0.0.0', port=5000)
